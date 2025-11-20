from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from django.db.models import Count, Sum, Q

# Import models
from users.models import Member
from books.models import Book, BookCopy
from loans.models import Loan

# Import untuk PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Import untuk Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io

from .utils import (
    get_current_academic_year,
    parse_academic_year,
    get_month_year_from_academic_year
)

@login_required
def reports_dashboard(request):
    """
    Dashboard untuk memilih jenis laporan
    """
    context = {
        'page_title': 'Laporan'
    }
    return render(request, 'reports/dashboard.html', context)


# ========== LAPORAN PEMINJAMAN ==========

@login_required
def loan_report_pdf(request):
    """
    Generate Laporan Peminjaman dalam format PDF
    """
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status', '')
    
    # Query loans
    loans = Loan.objects.select_related('member', 'book_copy__book').all()
    
    # Apply filters
    if start_date:
        loans = loans.filter(borrowed_date__gte=start_date)
    if end_date:
        loans = loans.filter(borrowed_date__lte=end_date)
    if status:
        loans = loans.filter(status=status)
    
    loans = loans.order_by('-borrowed_date')
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph("LAPORAN PEMINJAMAN BUKU", title_style)
    elements.append(title)
    
    # Info
    info_style = styles['Normal']
    if start_date and end_date:
        period = Paragraph(f"Periode: {start_date} s/d {end_date}", info_style)
        elements.append(period)
    
    generated = Paragraph(f"Digenerate: {timezone.now().strftime('%d %B %Y %H:%M')}", info_style)
    elements.append(generated)
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['No', 'Anggota', 'Buku', 'Tgl Pinjam', 'Jatuh Tempo', 'Status', 'Denda']]
    
    for idx, loan in enumerate(loans, 1):
        data.append([
            str(idx),
            f"{loan.member.name}\n{loan.member.nis}",
            loan.book_copy.book.title[:30],
            loan.borrowed_date.strftime('%d/%m/%Y'),
            loan.due_date.strftime('%d/%m/%Y'),
            loan.get_status_display(),
            f"Rp {loan.fine_amount:,.0f}" if loan.fine_amount > 0 else '-'
        ])
    
    # Create table
    table = Table(data, colWidths=[0.5*inch, 1.5*inch, 2*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Summary
    elements.append(Spacer(1, 20))
    total_loans = loans.count()
    total_fines = loans.aggregate(total=Sum('fine_amount'))['total'] or 0
    
    summary_data = [
        ['Total Peminjaman:', str(total_loans)],
        ['Total Denda:', f"Rp {total_fines:,.0f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"laporan_peminjaman_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def loan_report_excel(request):
    """
    Generate Laporan Peminjaman dalam format Excel
    """
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status', '')
    
    # Query loans
    loans = Loan.objects.select_related('member', 'book_copy__book').all()
    
    # Apply filters
    if start_date:
        loans = loans.filter(borrowed_date__gte=start_date)
    if end_date:
        loans = loans.filter(borrowed_date__lte=end_date)
    if status:
        loans = loans.filter(status=status)
    
    loans = loans.order_by('-borrowed_date')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Peminjaman"
    
    # Styles
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'LAPORAN PEMINJAMAN BUKU'
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Info
    ws['A2'] = f"Periode: {start_date or 'Semua'} s/d {end_date or 'Semua'}"
    ws['A3'] = f"Digenerate: {timezone.now().strftime('%d %B %Y %H:%M')}"
    
    # Headers
    headers = ['No', 'Anggota', 'NIS', 'Buku', 'Tgl Pinjam', 'Jatuh Tempo', 'Tgl Kembali', 'Status', 'Denda']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data
    row = 6
    total_fines = 0
    
    for idx, loan in enumerate(loans, 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=loan.member.name).border = border
        ws.cell(row=row, column=3, value=loan.member.nis).border = border
        ws.cell(row=row, column=4, value=loan.book_copy.book.title).border = border
        ws.cell(row=row, column=5, value=loan.borrowed_date.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=6, value=loan.due_date.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=7, value=loan.return_date.strftime('%d/%m/%Y') if loan.return_date else '-').border = border
        ws.cell(row=row, column=8, value=loan.get_status_display()).border = border
        ws.cell(row=row, column=9, value=loan.fine_amount).border = border
        
        total_fines += loan.fine_amount
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=8, value='Total Denda:').font = Font(bold=True)
    ws.cell(row=row, column=9, value=total_fines).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"laporan_peminjaman_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# ========== LAPORAN STATISTIK BULANAN (TAHUN AJARAN) ==========

@login_required
def monthly_report_pdf(request):
    """
    Generate Laporan Statistik Bulanan dalam format PDF (Tahun Ajaran)
    """
    # Get parameters
    month = int(request.GET.get('month', timezone.now().month))
    academic_year = request.GET.get('academic_year', '')
    
    # Parse academic year
    if academic_year:
        year_start, year_end = parse_academic_year(academic_year)
    else:
        year_start, year_end = get_current_academic_year()
    
    # Get actual year based on month
    year = get_month_year_from_academic_year(month, f"{year_start}/{year_end}")
    
    # Determine which year to use based on month
    # Juli (7) - Desember (12) = year_start
    # Januari (1) - Juni (6) = year_end
    if month >= 7:
        year = year_start
    else:
        year = year_end
    
    # Calculate date range
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1) - timedelta(days=1)
    
    # Get statistics
    total_loans = Loan.objects.filter(
        borrowed_date__gte=start_date,
        borrowed_date__lte=end_date
    ).count()
    
    total_returns = Loan.objects.filter(
        return_date__gte=start_date,
        return_date__lte=end_date
    ).count()
    
    total_fines = Loan.objects.filter(
        return_date__gte=start_date,
        return_date__lte=end_date
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    
    overdue_count = Loan.objects.filter(
        borrowed_date__gte=start_date,
        borrowed_date__lte=end_date,
        status='terlambat'
    ).count()
    
    active_loans = Loan.objects.filter(
        borrowed_date__gte=start_date,
        borrowed_date__lte=end_date,
        status__in=['dipinjam', 'terlambat']
    ).count()
    
    # Most borrowed books
    popular_books = Book.objects.annotate(
        loan_count=Count('bookcopy__loan', filter=Q(
            bookcopy__loan__borrowed_date__gte=start_date,
            bookcopy__loan__borrowed_date__lte=end_date
        ))
    ).filter(loan_count__gt=0).order_by('-loan_count')[:5]
    
    # Most active members
    active_members = Member.objects.annotate(
        loan_count=Count('loan', filter=Q(
            loan__borrowed_date__gte=start_date,
            loan__borrowed_date__lte=end_date
        ))
    ).filter(loan_count__gt=0).order_by('-loan_count')[:5]
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Title
    month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                   'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    
    title = Paragraph("LAPORAN STATISTIK BULANAN", title_style)
    elements.append(title)
    
    subtitle = Paragraph(f"Tahun Ajaran {year_start}/{year_end}<br/>Bulan: {month_names[month]} {year}", subtitle_style)
    elements.append(subtitle)
    
    # Info
    info_text = f"Digenerate: {timezone.now().strftime('%d %B %Y, %H:%M WIB')}"
    info = Paragraph(info_text, styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 20))
    
    # Statistics Summary
    stats_data = [
        ['STATISTIK PEMINJAMAN', 'JUMLAH'],
        ['Total Peminjaman', str(total_loans)],
        ['Total Pengembalian', str(total_returns)],
        ['Peminjaman Aktif', str(active_loans)],
        ['Peminjaman Terlambat', str(overdue_count)],
        ['Total Denda', f"Rp {total_fines:,.0f}"],
    ]
    
    stats_table = Table(stats_data, colWidths=[3.5*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 25))
    
    # Popular Books
    if popular_books.exists():
        popular_title = Paragraph("BUKU PALING BANYAK DIPINJAM", styles['Heading2'])
        elements.append(popular_title)
        elements.append(Spacer(1, 10))
        
        popular_data = [['No', 'Judul Buku', 'Penulis', 'Kategori', 'Jumlah']]
        
        for idx, book in enumerate(popular_books, 1):
            popular_data.append([
                str(idx),
                book.title[:35],
                book.author[:25],
                book.get_category_display(),
                str(book.loan_count)
            ])
        
        popular_table = Table(popular_data, colWidths=[0.4*inch, 2.5*inch, 1.8*inch, 1*inch, 0.8*inch])
        popular_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ]))
        
        elements.append(popular_table)
        elements.append(Spacer(1, 25))
    
    # Active Members
    if active_members.exists():
        members_title = Paragraph("ANGGOTA PALING AKTIF", styles['Heading2'])
        elements.append(members_title)
        elements.append(Spacer(1, 10))
        
        members_data = [['No', 'Nama Anggota', 'NIS/NIP', 'Tipe', 'Jumlah Pinjam']]
        
        for idx, member in enumerate(active_members, 1):
            members_data.append([
                str(idx),
                member.name[:30],
                member.nis,
                member.get_member_type_display(),
                str(member.loan_count)
            ])
        
        members_table = Table(members_data, colWidths=[0.4*inch, 2.5*inch, 1.2*inch, 1*inch, 1.2*inch])
        members_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ]))
        
        elements.append(members_table)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"laporan_bulanan_{year_start}_{year_end}_{month_names[month]}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def monthly_report_excel(request):
    """
    Generate Laporan Statistik Bulanan dalam format Excel (Tahun Ajaran)
    """
    # Get parameters
    month = int(request.GET.get('month', timezone.now().month))
    academic_year = request.GET.get('academic_year', '')
    
    # Parse academic year
    if academic_year:
        year_start, year_end = academic_year.split('/')
        year_start = int(year_start)
        year_end = int(year_end)
    else:
        current_year = timezone.now().year
        current_month = timezone.now().month
        if current_month >= 7:
            year_start = current_year
            year_end = current_year + 1
        else:
            year_start = current_year - 1
            year_end = current_year
    
    # Determine year based on month
    if month >= 7:
        year = year_start
    else:
        year = year_end
    
    # Calculate date range
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1) - timedelta(days=1)
    
    # Get statistics
    total_loans = Loan.objects.filter(
        borrowed_date__gte=start_date,
        borrowed_date__lte=end_date
    ).count()
    
    total_returns = Loan.objects.filter(
        return_date__gte=start_date,
        return_date__lte=end_date
    ).count()
    
    total_fines = Loan.objects.filter(
        return_date__gte=start_date,
        return_date__lte=end_date
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    
    overdue_count = Loan.objects.filter(
        borrowed_date__gte=start_date,
        borrowed_date__lte=end_date,
        status='terlambat'
    ).count()
    
    active_loans = Loan.objects.filter(
        borrowed_date__gte=start_date,
        borrowed_date__lte=end_date,
        status__in=['dipinjam', 'terlambat']
    ).count()
    
    # Most borrowed books
    popular_books = Book.objects.annotate(
        loan_count=Count('bookcopy__loan', filter=Q(
            bookcopy__loan__borrowed_date__gte=start_date,
            bookcopy__loan__borrowed_date__lte=end_date
        ))
    ).filter(loan_count__gt=0).order_by('-loan_count')[:10]
    
    # Most active members
    active_members = Member.objects.annotate(
        loan_count=Count('loan', filter=Q(
            loan__borrowed_date__gte=start_date,
            loan__borrowed_date__lte=end_date
        ))
    ).filter(loan_count__gt=0).order_by('-loan_count')[:10]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Bulanan"
    
    # Styles
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="667eea")
    
    # Title
    month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                   'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    
    ws.merge_cells('A1:E1')
    ws['A1'] = 'LAPORAN STATISTIK BULANAN'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f'Tahun Ajaran {year_start}/{year_end} - Bulan: {month_names[month]} {year}'
    ws['A2'].font = Font(bold=True, size=12, color="667eea")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = f"Digenerate: {timezone.now().strftime('%d %B %Y, %H:%M WIB')}"
    
    # Statistics
    ws['A5'] = 'STATISTIK PEMINJAMAN'
    ws['A5'].font = header_font
    ws['A5'].fill = header_fill
    ws['B5'] = 'JUMLAH'
    ws['B5'].font = header_font
    ws['B5'].fill = header_fill
    
    ws['A6'] = 'Total Peminjaman'
    ws['B6'] = total_loans
    ws['A7'] = 'Total Pengembalian'
    ws['B7'] = total_returns
    ws['A8'] = 'Peminjaman Aktif'
    ws['B8'] = active_loans
    ws['A9'] = 'Peminjaman Terlambat'
    ws['B9'] = overdue_count
    ws['A10'] = 'Total Denda'
    ws['B10'] = total_fines
    
    # Popular Books
    ws['A12'] = 'BUKU PALING BANYAK DIPINJAM'
    ws['A12'].font = Font(bold=True, size=14)
    
    headers = ['No', 'Judul Buku', 'Penulis', 'Kategori', 'Jumlah Peminjaman']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    row = 14
    for idx, book in enumerate(popular_books, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=book.title)
        ws.cell(row=row, column=3, value=book.author)
        ws.cell(row=row, column=4, value=book.get_category_display())
        ws.cell(row=row, column=5, value=book.loan_count)
        row += 1
    
    # Active Members
    row += 2
    ws.cell(row=row, column=1, value='ANGGOTA PALING AKTIF').font = Font(bold=True, size=14)
    
    row += 1
    headers = ['No', 'Nama Anggota', 'NIS/NIP', 'Tipe', 'Jumlah Peminjaman']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for idx, member in enumerate(active_members, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=member.name)
        ws.cell(row=row, column=3, value=member.nis)
        ws.cell(row=row, column=4, value=member.get_member_type_display())
        ws.cell(row=row, column=5, value=member.loan_count)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"laporan_bulanan_{year_start}_{year_end}_{month_names[month]}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ========== LAPORAN ANGGOTA ==========

@login_required
def member_report_pdf(request):
    """
    Generate Laporan Anggota dalam format PDF
    """
    members = Member.objects.filter(is_active=True).order_by('member_type', 'name')
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph("LAPORAN DATA ANGGOTA PERPUSTAKAAN", title_style)
    elements.append(title)
    
    # Info
    generated = Paragraph(f"Digenerate: {timezone.now().strftime('%d %B %Y %H:%M')}", styles['Normal'])
    elements.append(generated)
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['No', 'Nama', 'NIS/NIP', 'Tipe', 'Telepon', 'Status']]
    
    for idx, member in enumerate(members, 1):
        data.append([
            str(idx),
            member.name[:30],
            member.nis,
            member.get_member_type_display(),
            member.phone,
            'Aktif' if member.is_active else 'Nonaktif'
        ])
    
    # Create table
    table = Table(data, colWidths=[0.5*inch, 2*inch, 1.2*inch, 1*inch, 1.2*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Summary
    elements.append(Spacer(1, 20))
    summary = Paragraph(f"<b>Total Anggota Aktif: {members.count()}</b>", styles['Normal'])
    elements.append(summary)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"laporan_anggota_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def member_report_excel(request):
    """
    Generate Laporan Anggota dalam format Excel
    """
    members = Member.objects.filter(is_active=True).order_by('member_type', 'name')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Anggota"
    
    # Styles
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'LAPORAN DATA ANGGOTA PERPUSTAKAAN'
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Digenerate: {timezone.now().strftime('%d %B %Y %H:%M')}"
    
    # Headers
    headers = ['No', 'Nama', 'NIS/NIP', 'Tipe', 'Telepon', 'Email', 'Alamat']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    for idx, member in enumerate(members, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=member.name)
        ws.cell(row=row, column=3, value=member.nis)
        ws.cell(row=row, column=4, value=member.get_member_type_display())
        ws.cell(row=row, column=5, value=member.phone)
        ws.cell(row=row, column=6, value=member.email or '-')
        ws.cell(row=row, column=7, value=member.address)
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=1, value=f'Total Anggota Aktif: {members.count()}').font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 40
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"laporan_anggota_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ========== LAPORAN BUKU ==========

@login_required
def book_report_pdf(request):
    """
    Generate Laporan Buku dalam format PDF
    """
    books = Book.objects.all().order_by('category', 'title')
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph("LAPORAN DATA BUKU PERPUSTAKAAN", title_style)
    elements.append(title)
    
    # Info
    generated = Paragraph(f"Digenerate: {timezone.now().strftime('%d %B %Y %H:%M')}", styles['Normal'])
    elements.append(generated)
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['No', 'Judul', 'Penulis', 'Kategori', 'ISBN', 'Total', 'Tersedia']]
    
    for idx, book in enumerate(books, 1):
        data.append([
            str(idx),
            book.title[:35],
            book.author[:25],
            book.get_category_display(),
            book.isbn,
            str(book.total_copies),
            str(book.get_available_copies_count())
        ])
    
    # Create table
    table = Table(data, colWidths=[0.4*inch, 2.2*inch, 1.5*inch, 1*inch, 1.2*inch, 0.6*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))
    
    elements.append(table)
    
    # Summary
    elements.append(Spacer(1, 20))
    total_books = books.count()
    total_copies = sum([book.total_copies for book in books])
    total_available = sum([book.get_available_copies_count() for book in books])
    
    summary_data = [
        ['Total Judul Buku:', str(total_books)],
        ['Total Salinan:', str(total_copies)],
        ['Total Tersedia:', str(total_available)]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"laporan_buku_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def book_report_excel(request):
    """
    Generate Laporan Buku dalam format Excel
    """
    books = Book.objects.all().order_by('category', 'title')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Buku"
    
    # Styles
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'LAPORAN DATA BUKU PERPUSTAKAAN'
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Digenerate: {timezone.now().strftime('%d %B %Y %H:%M')}"
    
    # Headers
    headers = ['No', 'Judul', 'Penulis', 'Penerbit', 'Tahun', 'ISBN', 'Kategori', 'Total Salinan', 'Tersedia']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    for idx, book in enumerate(books, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=book.title)
        ws.cell(row=row, column=3, value=book.author)
        ws.cell(row=row, column=4, value=book.publisher)
        ws.cell(row=row, column=5, value=book.year_published)
        ws.cell(row=row, column=6, value=book.isbn)
        ws.cell(row=row, column=7, value=book.get_category_display())
        ws.cell(row=row, column=8, value=book.total_copies)
        ws.cell(row=row, column=9, value=book.get_available_copies_count())
        row += 1
    
    # Summary
    row += 1
    total_books = books.count()
    total_copies = sum([book.total_copies for book in books])
    total_available = sum([book.get_available_copies_count() for book in books])
    
    ws.cell(row=row, column=1, value='RINGKASAN').font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='Total Judul Buku:')
    ws.cell(row=row, column=2, value=total_books).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='Total Salinan:')
    ws.cell(row=row, column=2, value=total_copies).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='Total Tersedia:')
    ws.cell(row=row, column=2, value=total_available).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"laporan_buku_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ========== LAPORAN DENDA ==========

@login_required
def fine_report_pdf(request):
    """
    Generate Laporan Denda dalam format PDF
    """
    # Get loans with fines
    loans_with_fines = Loan.objects.filter(
        fine_amount__gt=0
    ).select_related('member', 'book_copy__book').order_by('-fine_amount')
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph("LAPORAN DENDA KETERLAMBATAN", title_style)
    elements.append(title)
    
    # Info
    generated = Paragraph(f"Digenerate: {timezone.now().strftime('%d %B %Y %H:%M')}", styles['Normal'])
    elements.append(generated)
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['No', 'Anggota', 'Buku', 'Tgl Pinjam', 'Tgl Kembali', 'Terlambat', 'Denda']]
    
    for idx, loan in enumerate(loans_with_fines, 1):
        days_late = (loan.return_date - loan.due_date).days if loan.return_date else 0
        data.append([
            str(idx),
            f"{loan.member.name}\n{loan.member.nis}",
            loan.book_copy.book.title[:30],
            loan.borrowed_date.strftime('%d/%m/%Y'),
            loan.return_date.strftime('%d/%m/%Y') if loan.return_date else '-',
            f"{days_late} hari",
            f"Rp {loan.fine_amount:,.0f}"
        ])
    
    # Create table
    table = Table(data, colWidths=[0.4*inch, 1.5*inch, 2*inch, 0.9*inch, 0.9*inch, 0.8*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))
    
    elements.append(table)
    
    # Summary
    elements.append(Spacer(1, 20))
    total_fines = loans_with_fines.aggregate(total=Sum('fine_amount'))['total'] or 0
    
    summary = Paragraph(f"<b>Total Denda: Rp {total_fines:,.0f}</b>", styles['Heading3'])
    elements.append(summary)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"laporan_denda_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def fine_report_excel(request):
    """
    Generate Laporan Denda dalam format Excel
    """
    # Get loans with fines
    loans_with_fines = Loan.objects.filter(
        fine_amount__gt=0
    ).select_related('member', 'book_copy__book').order_by('-fine_amount')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Denda"
    
    # Styles
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'LAPORAN DENDA KETERLAMBATAN'
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Digenerate: {timezone.now().strftime('%d %B %Y %H:%M')}"
    
    # Headers
    headers = ['No', 'Anggota', 'NIS', 'Buku', 'Tgl Pinjam', 'Tgl Kembali', 'Terlambat (hari)', 'Denda']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    for idx, loan in enumerate(loans_with_fines, 1):
        days_late = (loan.return_date - loan.due_date).days if loan.return_date else 0
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=loan.member.name)
        ws.cell(row=row, column=3, value=loan.member.nis)
        ws.cell(row=row, column=4, value=loan.book_copy.book.title)
        ws.cell(row=row, column=5, value=loan.borrowed_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=6, value=loan.return_date.strftime('%d/%m/%Y') if loan.return_date else '-')
        ws.cell(row=row, column=7, value=days_late)
        ws.cell(row=row, column=8, value=loan.fine_amount)
        row += 1
    
    # Summary
    row += 1
    total_fines = loans_with_fines.aggregate(total=Sum('fine_amount'))['total'] or 0
    ws.cell(row=row, column=7, value='TOTAL DENDA:').font = Font(bold=True)
    ws.cell(row=row, column=8, value=total_fines).font = Font(bold=True, size=12)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"laporan_denda_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
    
    